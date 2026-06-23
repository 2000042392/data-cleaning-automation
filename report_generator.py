"""report_generator.py — Charts, Excel, and HTML from clean data."""
import base64, warnings
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
warnings.filterwarnings("ignore")

PALETTE = ["#2D6A9F","#3AA6B9","#74C69D","#F4A261","#E76F51","#9B72CF"]
BG,DARK,ACCENT = "#F8FAFC","#1E293B","#2D6A9F"
sns.set_theme(style="whitegrid", palette=PALETTE)
plt.rcParams.update({"figure.facecolor":BG,"axes.facecolor":BG,"axes.edgecolor":"#CBD5E1",
    "axes.titlesize":14,"axes.titleweight":"bold","axes.titlecolor":DARK,
    "axes.labelcolor":DARK,"text.color":DARK,"font.family":"DejaVu Sans",
    "grid.color":"#E2E8F0","grid.linewidth":0.7})


class ReportGenerator:
    def __init__(self, df, out_dir="output"):
        self.df = df.copy()
        self.charts_dir = Path(out_dir)/"charts"
        self.excel_dir  = Path(out_dir)/"excel"
        self.report_dir = Path(out_dir)/"reports"
        for d in [self.charts_dir,self.excel_dir,self.report_dir]:
            d.mkdir(parents=True,exist_ok=True)
        self.chart_paths = {}

    def _save(self, fig, name):
        path = self.charts_dir/f"{name}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=BG)
        plt.close(fig)
        self.chart_paths[name] = str(path)

    def chart_revenue_by_region(self):
        fig,ax = plt.subplots(figsize=(9,5))
        d = self.df.groupby("region")["net_revenue"].sum().sort_values(ascending=False).reset_index()
        bars = ax.bar(d["region"],d["net_revenue"],color=PALETTE[:len(d)],edgecolor="white",linewidth=1.2,zorder=3)
        for b in bars:
            h=b.get_height()
            ax.text(b.get_x()+b.get_width()/2,h+800,f"${h:,.0f}",ha="center",va="bottom",fontsize=10,fontweight="bold")
        ax.set_title("Net Revenue by Region",pad=15); ax.set_ylabel("Net Revenue (USD)")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"${x:,.0f}"))
        ax.set_ylim(0,d["net_revenue"].max()*1.18); ax.yaxis.grid(True,zorder=0); ax.set_axisbelow(True)
        plt.tight_layout(); self._save(fig,"revenue_by_region")

    def chart_monthly_trend(self):
        fig,ax = plt.subplots(figsize=(11,5))
        m = self.df.groupby(self.df["date"].dt.to_period("M"))["net_revenue"].sum().reset_index()
        m["date"] = m["date"].astype(str)
        ax.plot(m["date"],m["net_revenue"],color=ACCENT,linewidth=2.5,marker="o",markersize=6,
                markerfacecolor="white",markeredgewidth=2,zorder=3)
        ax.fill_between(range(len(m)),m["net_revenue"],alpha=0.12,color=ACCENT)
        ax.set_xticks(range(len(m))); ax.set_xticklabels(m["date"],rotation=45,ha="right",fontsize=9)
        ax.set_title("Monthly Net Revenue Trend",pad=15); ax.set_ylabel("Net Revenue (USD)")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"${x:,.0f}"))
        ax.yaxis.grid(True,zorder=0); ax.set_axisbelow(True)
        plt.tight_layout(); self._save(fig,"monthly_trend")

    def chart_product_mix(self):
        fig,axes = plt.subplots(1,2,figsize=(12,5))
        prod = self.df.groupby("product")["net_revenue"].sum().sort_values(ascending=False)
        axes[0].pie(prod.values,labels=prod.index,autopct="%1.1f%%",colors=PALETTE[:len(prod)],
                    startangle=140,wedgeprops=dict(width=0.55,edgecolor="white",linewidth=2),
                    pctdistance=0.75,textprops={"fontsize":9})
        axes[0].set_title("Revenue Share by Product")
        axes[1].barh(prod.index[::-1],prod.values[::-1],color=PALETTE[:len(prod)],edgecolor="white",linewidth=1)
        for i,v in enumerate(prod.values[::-1]):
            axes[1].text(v+200,i,f"${v:,.0f}",va="center",fontsize=9)
        axes[1].set_xlabel("Net Revenue (USD)"); axes[1].set_title("Revenue by Product")
        axes[1].xaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"${x:,.0f}"))
        plt.tight_layout(); self._save(fig,"product_mix")

    def chart_status_distribution(self):
        fig,ax = plt.subplots(figsize=(8,5))
        s = self.df["status"].value_counts()
        bars = ax.bar(s.index,s.values,color=PALETTE[:len(s)],edgecolor="white",linewidth=1.2)
        for b in bars:
            h=b.get_height()
            ax.text(b.get_x()+b.get_width()/2,h+1,str(int(h)),ha="center",va="bottom",fontsize=11,fontweight="bold")
        ax.set_title("Orders by Status"); ax.set_ylabel("Number of Orders")
        ax.yaxis.grid(True,zorder=0); ax.set_axisbelow(True)
        plt.tight_layout(); self._save(fig,"status_distribution")

    def chart_top_reps(self):
        fig,ax = plt.subplots(figsize=(9,5))
        reps = self.df.groupby("sales_rep")["net_revenue"].sum().sort_values(ascending=False).head(8)
        colors = [PALETTE[0] if i==0 else PALETTE[2] for i in range(len(reps))]
        ax.barh(reps.index[::-1],reps.values[::-1],color=colors[::-1],edgecolor="white",linewidth=1.2)
        for i,v in enumerate(reps.values[::-1]):
            ax.text(v+200,i,f"${v:,.0f}",va="center",fontsize=9)
        ax.set_xlabel("Net Revenue (USD)"); ax.set_title("Top Sales Representatives")
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"${x:,.0f}"))
        gold = mpatches.Patch(color=PALETTE[0],label="Top performer")
        ax.legend(handles=[gold],loc="lower right")
        plt.tight_layout(); self._save(fig,"top_reps")

    def generate_all_charts(self):
        self.chart_revenue_by_region(); self.chart_monthly_trend()
        self.chart_product_mix(); self.chart_status_distribution(); self.chart_top_reps()
        print(f"✓ Generated {len(self.chart_paths)} charts")

    def generate_excel(self):
        wb = Workbook()
        HDR  = PatternFill("solid",fgColor="2D6A9F")
        ALT  = PatternFill("solid",fgColor="EFF6FF")
        TOT  = PatternFill("solid",fgColor="DBEAFE")
        WHT  = PatternFill("solid",fgColor="FFFFFF")
        thin = Side(style="thin",color="CBD5E1")
        bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)

        def hdr(cell,val):
            cell.value=val; cell.font=Font(bold=True,color="FFFFFF",size=11,name="Arial")
            cell.fill=HDR; cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=bdr

        def dat(cell,val,fmt=None,bold=False,fill=None):
            cell.value=val; cell.font=Font(bold=bold,size=10,name="Arial",color="2D6A9F" if bold else "1E293B")
            cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.border=bdr
            if fmt: cell.number_format=fmt
            cell.fill=fill if fill else WHT

        # Sheet 1 — Cleaned Data
        ws1=wb.active; ws1.title="Cleaned Data"; ws1.sheet_view.showGridLines=False; ws1.freeze_panes="A2"
        cols=list(self.df.columns)
        for ci,c in enumerate(cols,1): hdr(ws1.cell(1,ci),c.replace("_"," ").title())
        for ri,row in enumerate(self.df.itertuples(index=False),2):
            fill=ALT if ri%2==0 else WHT
            for ci,val in enumerate(row,1):
                col=cols[ci-1]; fmt=None
                if col in("revenue","net_revenue"): fmt="#,##0.00"
                elif col=="discount_%": fmt="0.0"
                dat(ws1.cell(ri,ci),val,fmt=fmt,fill=fill)
        for ci,col in enumerate(cols,1):
            maxl=max(len(col),*[len(str(ws1.cell(r,ci).value or "")) for r in range(2,min(20,ws1.max_row+1))])
            ws1.column_dimensions[get_column_letter(ci)].width=min(maxl+4,24)
        ws1.row_dimensions[1].height=32

        # Sheet 2 — Regional Summary
        ws2=wb.create_sheet("Regional Summary"); ws2.sheet_view.showGridLines=False
        rdata=self.df.groupby("region").agg(orders=("order_id","count"),total_revenue=("revenue","sum"),
            net_revenue=("net_revenue","sum"),avg_order_value=("net_revenue","mean"),
            avg_rating=("customer_rating","mean")).round(2).reset_index().sort_values("net_revenue",ascending=False)
        for ci,h in enumerate(["Region","Orders","Total Revenue","Net Revenue","Avg Order Value","Avg Rating"],1): hdr(ws2.cell(1,ci),h)
        for ri,row in enumerate(rdata.itertuples(index=False),2):
            fill=ALT if ri%2==0 else WHT; vals=list(row)
            fmts=[None,"#,##0","#,##0.00","#,##0.00","#,##0.00","0.00"]
            for ci,(v,f) in enumerate(zip(vals,fmts),1): dat(ws2.cell(ri,ci),v,fmt=f,fill=fill)
        tr=len(rdata)+2; dat(ws2.cell(tr,1),"TOTAL",bold=True,fill=TOT)
        for ci in range(2,7):
            cl=get_column_letter(ci); c=ws2.cell(tr,ci)
            c.value=f"=SUM({cl}2:{cl}{tr-1})" if ci in(2,3,4) else f"=AVERAGE({cl}2:{cl}{tr-1})"
            c.font=Font(bold=True,name="Arial",size=10,color="2D6A9F"); c.fill=TOT
            c.alignment=Alignment(horizontal="center"); c.border=bdr
            c.number_format=["","#,##0","#,##0.00","#,##0.00","#,##0.00","0.00"][ci-1]
        for ci in range(1,7): ws2.column_dimensions[get_column_letter(ci)].width=20
        ws2.row_dimensions[1].height=30

        # Sheet 3 — Product Performance
        ws3=wb.create_sheet("Product Performance"); ws3.sheet_view.showGridLines=False
        pdata=self.df.groupby("product").agg(orders=("order_id","count"),units_sold=("quantity","sum"),
            net_revenue=("net_revenue","sum"),avg_discount=("discount_%","mean")).round(2).reset_index()
        pdata["revenue_per_unit"]=(pdata["net_revenue"]/pdata["units_sold"]).round(2)
        pdata=pdata.sort_values("net_revenue",ascending=False)
        for ci,h in enumerate(["Product","Orders","Units Sold","Net Revenue","Avg Discount %","Revenue / Unit"],1): hdr(ws3.cell(1,ci),h)
        for ri,row in enumerate(pdata.itertuples(index=False),2):
            fill=ALT if ri%2==0 else WHT; vals=list(row)
            fmts=[None,"#,##0","#,##0","#,##0.00","0.0","#,##0.00"]
            for ci,(v,f) in enumerate(zip(vals,fmts),1): dat(ws3.cell(ri,ci),v,fmt=f,fill=fill)
        for ci in range(1,7): ws3.column_dimensions[get_column_letter(ci)].width=20
        ws3.row_dimensions[1].height=30

        # Sheet 4 — Monthly Trend
        ws4=wb.create_sheet("Monthly Trend"); ws4.sheet_view.showGridLines=False
        monthly=self.df.groupby(self.df["date"].dt.to_period("M")).agg(
            orders=("order_id","count"),net_revenue=("net_revenue","sum")).reset_index()
        monthly["date"]=monthly["date"].astype(str)
        for ci,h in enumerate(["Month","Orders","Net Revenue"],1): hdr(ws4.cell(1,ci),h)
        for ri,row in enumerate(monthly.itertuples(index=False),2):
            fill=ALT if ri%2==0 else WHT
            dat(ws4.cell(ri,1),row.date,fill=fill)
            dat(ws4.cell(ri,2),row.orders,fmt="#,##0",fill=fill)
            dat(ws4.cell(ri,3),row.net_revenue,fmt="#,##0.00",fill=fill)
        for ci in range(1,4): ws4.column_dimensions[get_column_letter(ci)].width=20
        ws4.row_dimensions[1].height=30

        path=self.excel_dir/"sales_report.xlsx"; wb.save(str(path))
        print(f"✓ Excel workbook → {path}"); return str(path)

    def generate_html(self, summary, audit):
        def b64(p):
            with open(p,"rb") as f: return "data:image/png;base64,"+base64.b64encode(f.read()).decode()
        imgs={k:b64(v) for k,v in self.chart_paths.items()}
        kpis={"total_rev":f"${self.df['net_revenue'].sum():,.0f}",
              "orders":f"{len(self.df):,}","avg_order":f"${self.df['net_revenue'].mean():,.0f}",
              "top_region":self.df.groupby("region")["net_revenue"].sum().idxmax(),
              "top_prod":self.df.groupby("product")["net_revenue"].sum().idxmax(),
              "rating":f"{self.df['customer_rating'].mean():.2f}"}
        audit_rows="".join(
            f"<tr><td>{r.step}</td><td>{r.detail}</td><td class='num'>{r.rows_affected}</td><td>{r.timestamp}</td></tr>"
            for r in audit.itertuples(index=False))

        html=f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Data Cleaning & Reporting — Automated Report</title>
<style>
:root{{--bg:#F8FAFC;--card:#fff;--dark:#1E293B;--mid:#475569;--light:#94A3B8;
      --accent:#2D6A9F;--a2:#3AA6B9;--green:#10B981;--orange:#F4A261;
      --border:#E2E8F0;--r:12px}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:var(--bg);color:var(--dark)}}
.topbar{{background:linear-gradient(135deg,#2D6A9F,#3AA6B9);color:#fff;padding:32px 48px}}
.topbar h1{{font-size:2rem;font-weight:700;letter-spacing:-0.5px}}
.topbar p{{margin-top:6px;opacity:.85;font-size:.95rem}}
.badge{{display:inline-block;background:rgba(255,255,255,.2);border-radius:50px;
        padding:4px 14px;font-size:.8rem;margin-top:12px;letter-spacing:.5px}}
main{{max-width:1200px;margin:0 auto;padding:40px 24px}}
h2{{font-size:1.3rem;font-weight:700;color:var(--dark);margin-bottom:20px;
    display:flex;align-items:center;gap:8px}}
h2::before{{content:'';display:block;width:4px;height:22px;background:var(--accent);border-radius:2px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;margin-bottom:40px}}
.kpi{{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:20px;
      text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.05);transition:.2s}}
.kpi:hover{{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,.08)}}
.kpi .val{{font-size:1.8rem;font-weight:800;color:var(--accent);line-height:1}}
.kpi .lbl{{font-size:.78rem;color:var(--light);margin-top:6px;text-transform:uppercase;letter-spacing:.6px}}
.kpi.green .val{{color:var(--green)}}.kpi.orange .val{{color:var(--orange)}}
.clean-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:14px;margin-bottom:40px}}
.cc{{background:var(--card);border:1px solid var(--border);border-radius:var(--r);
     padding:18px 16px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.04)}}
.cc .num{{font-size:2rem;font-weight:800;color:var(--a2)}}.cc .desc{{font-size:.8rem;color:var(--mid);margin-top:4px}}
.chart-grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:40px}}
.chart-wide{{grid-column:1/-1}}
.chart-box{{background:var(--card);border:1px solid var(--border);border-radius:var(--r);
            padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.05)}}
.chart-box img{{width:100%;border-radius:8px}}
.table-wrap{{overflow-x:auto;margin-bottom:40px}}
table{{width:100%;border-collapse:collapse;font-size:.85rem}}
thead th{{background:var(--accent);color:#fff;padding:10px 14px;text-align:left;font-weight:600;white-space:nowrap}}
tbody tr:nth-child(even){{background:#EFF6FF}}
tbody td{{padding:9px 14px;border-bottom:1px solid var(--border);color:var(--mid)}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
footer{{text-align:center;padding:32px;color:var(--light);font-size:.82rem;
        border-top:1px solid var(--border);margin-top:20px}}
@media(max-width:700px){{.chart-grid{{grid-template-columns:1fr}}}}
</style></head><body>
<div class="topbar">
  <h1>📊 Data Cleaning &amp; Reporting Automation</h1>
  <p>Fully automated pipeline — raw data ingested, cleaned, and summarised without manual intervention</p>
  <span class="badge">Generated automatically · Python · Pandas · Matplotlib · OpenPyXL</span>
</div>
<main>
<h2>Key Performance Indicators</h2>
<div class="kpi-grid">
  <div class="kpi"><div class="val">{kpis['total_rev']}</div><div class="lbl">Total Net Revenue</div></div>
  <div class="kpi"><div class="val">{kpis['orders']}</div><div class="lbl">Clean Orders</div></div>
  <div class="kpi green"><div class="val">{kpis['avg_order']}</div><div class="lbl">Avg Order Value</div></div>
  <div class="kpi orange"><div class="val">{kpis['top_region']}</div><div class="lbl">Top Region</div></div>
  <div class="kpi"><div class="val">{kpis['top_prod']}</div><div class="lbl">Top Product</div></div>
  <div class="kpi green"><div class="val">{kpis['rating']} ⭐</div><div class="lbl">Avg Customer Rating</div></div>
</div>
<h2>Data Cleaning Summary</h2>
<div class="clean-grid">
  <div class="cc"><div class="num">{summary['original_rows']}</div><div class="desc">Raw Rows Ingested</div></div>
  <div class="cc"><div class="num">{summary['clean_rows']}</div><div class="desc">Clean Rows Output</div></div>
  <div class="cc"><div class="num">{summary['rows_removed']}</div><div class="desc">Rows Removed</div></div>
  <div class="cc"><div class="num">{summary['original_duplicates']}</div><div class="desc">Duplicates Removed</div></div>
  <div class="cc"><div class="num">{summary['original_nulls']}</div><div class="desc">Nulls in Raw Data</div></div>
  <div class="cc"><div class="num">{summary['clean_nulls']}</div><div class="desc">Remaining Nulls</div></div>
</div>
<h2>Visual Summaries</h2>
<div class="chart-grid">
  <div class="chart-box chart-wide"><img src="{imgs.get('monthly_trend','')}" alt="Monthly Trend"></div>
  <div class="chart-box"><img src="{imgs.get('revenue_by_region','')}" alt="Revenue by Region"></div>
  <div class="chart-box"><img src="{imgs.get('product_mix','')}" alt="Product Mix"></div>
  <div class="chart-box"><img src="{imgs.get('status_distribution','')}" alt="Status Distribution"></div>
  <div class="chart-box"><img src="{imgs.get('top_reps','')}" alt="Top Reps"></div>
</div>
<h2>Cleaning Audit Trail</h2>
<div class="table-wrap">
<table><thead><tr><th>Step</th><th>Detail</th><th>Rows Affected</th><th>Time</th></tr></thead>
<tbody>{audit_rows}</tbody></table></div>
</main>
<footer>Auto-generated by <strong>data_pipeline.py</strong> · Data Cleaning &amp; Reporting Automation Project</footer>
</body></html>"""
        path=self.report_dir/"report.html"; path.write_text(html,encoding="utf-8")
        print(f"✓ HTML report → {path}"); return str(path)
